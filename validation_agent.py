import os
import io
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from playwright.sync_api import sync_playwright
from dotenv import load_dotenv
import psutil

# Load environment variables
load_dotenv()

# Configuration
SCOPES = ['https://www.googleapis.com/auth/drive.readonly']
CREDENTIALS_FILE = 'credentials.json'
TOKEN_FILE = 'token.json'
INPUT_FILE = 'Recovered_Policies.xlsx'
OUTPUT_FILE = 'Verified_Policies.xlsx'
PROGRESS_FILE = 'Verified_Policies_Progress.xlsx'
CHECKPOINT_INTERVAL = 500

# APL Portal Credentials
APL_COMPANY_ID = os.environ.get('APL_COMPANY_ID')
APL_CARRIER = os.environ.get('APL_CARRIER')
APL_USER = os.environ.get('APL_USER')
APL_PASS = os.environ.get('APL_PASS')
CHROME_USER_DATA_DIR = os.environ.get('CHROME_USER_DATA_DIR')
HEADLESS = os.getenv('HEADLESS', 'False').lower() == 'true'

# Carrier Name Aliases for Google Drive Folder Matching
# Add any carrier abbreviations or alternate names here. Keys should be lowercase.
CARRIER_ALIASES = {
    "independence blue cross": ["ibc"],
    "united healthcare": ["uhc", "unitedhealth"],
    "blue cross blue shield": ["bcbs"],
    "health care service corp": ["hcsc"]
}

def check_chrome_running():
    for proc in psutil.process_iter(['name']):
        if proc.info['name'] and 'chrome.exe' in proc.info['name'].lower():
            print("⚠️ WARNING: Google Chrome is currently running. This may cause Playwright to crash if using the same profile.")
            print("Please close Chrome or ensure you are using a dedicated profile for automation.")
            break

class ValidationAgent:
    def __init__(self):
        self.drive_service = None
        self.cached_statements = {} # Dictionary keyed by carrier name
        self.playwright = None
        self.browser_context = None
        self.page = None
        self.is_logged_in = False
        
    def _authenticate_drive(self):
        print("🔐 Authenticating Google Drive...")
        creds = None
        if os.path.exists(TOKEN_FILE):
            creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                if not os.path.exists(CREDENTIALS_FILE):
                    print(f"⚠️ ERROR: {CREDENTIALS_FILE} not found. Please setup Google Cloud Project OAuth.")
                    return None
                flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
                creds = flow.run_local_server(port=0)
            with open(TOKEN_FILE, 'w') as token:
                token.write(creds.to_json())
        return build('drive', 'v3', credentials=creds)

    def _get_all_excel_files_in_folder(self, folder_id):
        all_files = []
        page_token = None
        while True:
            query = f"'{folder_id}' in parents"
            
            # Add robust retry logic for network drops
            max_retries = 3
            results = None
            for attempt in range(max_retries):
                try:
                    results = self.drive_service.files().list(
                        q=query, corpora='allDrives', includeItemsFromAllDrives=True, 
                        supportsAllDrives=True, fields='nextPageToken, files(id, name, mimeType)',
                        pageToken=page_token).execute()
                    break
                except Exception as e:
                    if attempt == max_retries - 1:
                        raise e
                    print(f"      ⚠️ Network hiccup detected. Retrying API request ({attempt+1}/{max_retries})...")
                    time.sleep(2)
            
            
            items = results.get('files', [])
            for item in items:
                mime = item['mimeType']
                if mime == 'application/vnd.google-apps.folder':
                    all_files.extend(self._get_all_excel_files_in_folder(item['id']))
                elif mime == 'application/vnd.google-apps.spreadsheet' or mime == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
                    all_files.append(item)
                    
            page_token = results.get('nextPageToken')
            if not page_token:
                break
        return all_files

    def cache_carrier_statements(self, target_carriers=None):
        print("📂 Locating 'Carrier_Statements' folder...")
        if not self.drive_service:
            print("❌ Google Drive API is not authenticated.")
            return

        query = "name = 'Carrier_Statements' and mimeType = 'application/vnd.google-apps.folder'"
        results = self.drive_service.files().list(q=query, corpora='allDrives', includeItemsFromAllDrives=True, supportsAllDrives=True, fields='files(id, name)').execute()
        folders = results.get('files', [])
        
        if not folders:
            print("❌ Folder 'Carrier_Statements' not found in Google Drive.")
            return
            
        root_folder_id = folders[0]['id']
        print(f"✅ Found root folder ID: {root_folder_id}. Locating carrier subfolders...")
        
        # Get all top-level carrier subfolders with pagination
        subfolders = []
        page_token = None
        while True:
            query = f"'{root_folder_id}' in parents and mimeType='application/vnd.google-apps.folder'"
            results = self.drive_service.files().list(
                q=query, corpora='allDrives', includeItemsFromAllDrives=True, 
                supportsAllDrives=True, fields='nextPageToken, files(id, name)',
                pageToken=page_token).execute()
            subfolders.extend(results.get('files', []))
            page_token = results.get('nextPageToken')
            if not page_token:
                break

        print(f"📁 Found {len(subfolders)} top-level carrier subfolders.")
        
        # Track file IDs to prevent any possibility of infinite loops or duplicate downloads
        seen_file_ids = set()
        
        for subfolder in subfolders:
            sub_id = subfolder['id']
            sub_name = subfolder['name'].lower()
            
            # OPTIMIZATION: Only process folders that match our target carriers
            if target_carriers:
                is_target = any(tc in sub_name or sub_name in tc for tc in target_carriers)
                if not is_target:
                    continue
                    
            self.cached_statements[sub_name] = []
            
            print(f"   🔍 Recursively scanning folder: {subfolder['name']}")
            files = self._get_all_excel_files_in_folder(sub_id)
            
            for file in files:
                file_id = file['id']
                
                # Loop Protection: Strictly ignore if we have already downloaded this exact file ID
                if file_id in seen_file_ids:
                    continue
                seen_file_ids.add(file_id)
                
                name = file['name']
                mime_type = file['mimeType']
                name_lower = name.lower()
                
                if ".csv" in name_lower or "recovered policies" in name_lower or "policy list combined" in name_lower:
                    print(f"      ⏩ Skipping excluded file: {name}")
                    continue
                    
                try:
                    if mime_type == 'application/vnd.google-apps.spreadsheet':
                        request = self.drive_service.files().export_media(fileId=file_id, mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    else:
                        request = self.drive_service.files().get_media(fileId=file_id)
                        
                    fh = io.BytesIO()
                    downloader = MediaIoBaseDownload(fh, request)
                    done = False
                    while done is False:
                        # Retry logic for downloading chunks
                        max_download_retries = 3
                        for attempt in range(max_download_retries):
                            try:
                                status, done = downloader.next_chunk()
                                break
                            except Exception as e:
                                if attempt == max_download_retries - 1:
                                    raise e
                                print(f"      ⚠️ Download interrupted. Retrying chunk ({attempt+1}/{max_download_retries})...")
                                time.sleep(2)
                    
                    fh.seek(0)
                    df_dict = pd.read_excel(fh, sheet_name=None)
                    for sheet_name, df in df_dict.items():
                        text_dump = df.to_csv(index=False).lower()
                        self.cached_statements[sub_name].append((name, sheet_name, text_dump))
                    print(f"      - Cached: {name}")
                except Exception as e:
                    print(f"      ❌ Failed to cache {name}: {e}")

    def _local_search(self, policy_number, carrier_name):
        policy_lower = str(policy_number).lower()
        carrier_lower = str(carrier_name).lower()
        
        # Collect all valid aliases for the given carrier
        valid_names = [carrier_lower]
        if carrier_lower in CARRIER_ALIASES:
            valid_names.extend(CARRIER_ALIASES[carrier_lower])
        
        target_files = []
        for folder_name, files in self.cached_statements.items():
            # If any of the valid aliases match the folder name
            if any(name in folder_name or folder_name in name for name in valid_names):
                target_files.extend(files)
                
        for name, sheet_name, text_dump in target_files:
            if policy_lower in text_dump:
                return True
        return False

    def validate_in_drive(self, policy_number, carrier_name):
        if not policy_number or pd.isna(policy_number):
            return "No Match", "Policy number is empty"
            
        policy_number = str(policy_number).strip()
        
        # 1. Exact Match
        if self._local_search(policy_number, carrier_name):
            return "Exact Match", "Matched perfectly"
            
        # 2. Fuzzy Match
        # Strip from front
        stripped_front = policy_number[1:]
        if len(stripped_front) > 0 and self._local_search(stripped_front, carrier_name):
            return "Fuzzy Match", f"Matched after stripping 1 from front"
        
        # Strip from back
        stripped_back = policy_number[:-1]
        if len(stripped_back) > 0 and self._local_search(stripped_back, carrier_name):
            return "Fuzzy Match", f"Matched after stripping 1 from back"
            
        return "No Match", "Not found in any statements"

    def start_browser(self):
        if not self.playwright:
            self.playwright = sync_playwright().start()
            
        if self.browser_context:
            try:
                self.browser_context.close()
            except:
                pass
                
        self.browser_context = self.playwright.chromium.launch_persistent_context(
            user_data_dir=CHROME_USER_DATA_DIR,
            headless=HEADLESS,
            args=['--no-sandbox', '--disable-setuid-sandbox']
        )
        
        if len(self.browser_context.pages) > 0:
            self.page = self.browser_context.pages[0]
        else:
            self.page = self.browser_context.new_page()
            
        self.page.on("dialog", lambda dialog: dialog.accept())
        self.is_logged_in = False

    def close_browser(self):
        if self.browser_context:
            try:
                self.browser_context.close()
            except Exception:
                pass
            self.browser_context = None
            
        if self.playwright:
            try:
                self.playwright.stop()
            except Exception:
                pass
            self.playwright = None
            
        self.is_logged_in = False

    def login_to_apl(self):
        print("🌐 Logging into APL portal...")
        self.page.goto("https://p20.aplplus.com/", timeout=60000)
        
        # 1. Login Form (Using precise IDs from HTML)
        if self.page.locator("#txtCompanyID").is_visible():
            self.page.fill("#txtCompanyID", APL_COMPANY_ID or "")
            self.page.press("#txtCompanyID", "Tab")
            
            try:
                self.page.wait_for_load_state("networkidle", timeout=5000)
            except:
                pass
                
            if APL_CARRIER:
                try:
                    self.page.locator(f"#ddDB option:has-text('{APL_CARRIER}')").wait_for(state="attached", timeout=5000)
                except:
                    pass
                self.page.select_option("#ddDB", label=APL_CARRIER)
                
            self.page.fill("#txtUserID", APL_USER or "")
            self.page.fill("#txtpwd", APL_PASS or "")
            self.page.click("#iplogin")
            
            try:
                self.page.wait_for_load_state("networkidle", timeout=10000)
            except:
                pass
            
            # Handle "Password expires" custom modal
            try:
                ok_btn = self.page.locator("button:has-text('Ok'), input[value='Ok'], button:has-text('OK'), input[value='OK']").first
                ok_btn.wait_for(state="visible", timeout=5000)
                if ok_btn.is_visible():
                    ok_btn.click()
                    try:
                        self.page.wait_for_load_state("networkidle", timeout=5000)
                    except:
                        pass
            except:
                pass
                
        # 2. Navigate to Policies Tab
        try:
            self.page.get_by_text("Policies", exact=True).click(timeout=5000)
            try:
                self.page.wait_for_load_state("networkidle", timeout=5000)
            except:
                pass
        except:
            pass # Already on it or different layout
            
        self.is_logged_in = True

    def verify_apl_portal(self, policy_number, retry=True):
        print(f"🌐 Verifying {policy_number} on APL portal...")
        try:
            if not self.browser_context or not self.is_logged_in:
                self.start_browser()
                self.login_to_apl()
                
            # 3. Search for policy
            # Wait for the specific Quick Search input
            search_input = self.page.locator("input[placeholder*='Search']").first
            search_input.wait_for(state="visible", timeout=10000)
            search_input.fill(str(policy_number))
            search_input.press("Enter")
            
            try:
                self.page.wait_for_load_state("networkidle", timeout=10000)
            except:
                pass
            self.page.wait_for_timeout(2000) # Give table time to populate
            
            # 4. Check for EXACT match in the first column of the table
            policy_rows = self.page.locator("table tbody tr")
            count = policy_rows.count()
            for i in range(count):
                row = policy_rows.nth(i)
                first_cell = row.locator("td").nth(0)
                if first_cell.is_visible():
                    cell_text = first_cell.inner_text().strip()
                    if cell_text == str(policy_number):
                        return "APL Match", "Found exact match in APL portal"
            
            return "No Match", "Policy not found in APL portal"
                
        except Exception as e:
            print(f"⚠️ Playwright automation failed for {policy_number}: {e}")
            if retry:
                print("🔄 Portal stuck or errored. Destroying browser and retrying...")
                try:
                    self.page.screenshot(path=f"debug_screenshot_{policy_number}.png")
                except:
                    pass
                self.close_browser()
                return self.verify_apl_portal(policy_number, retry=False)
            return "Error", f"Automation failed: {str(e)}"

    def format_excel_output(self, file_path):
        print(f"🎨 Formatting Excel file: {file_path}")
        wb = load_workbook(file_path)
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="333F4F", fill_type="solid")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Style header row
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                # Cap the maximum width for readability
                ws.column_dimensions[column].width = min(adjusted_width, 50)
                
        wb.save(file_path)

    def run(self):
        check_chrome_running()
        
        print("🚀 Starting Validation Agent...")
        self.drive_service = self._authenticate_drive()
        
        # 1. Load Input Data FIRST
        input_data = {}
        if os.path.exists(PROGRESS_FILE):
            print(f"📄 Resuming from checkpoint {PROGRESS_FILE}...")
            input_data = pd.read_excel(PROGRESS_FILE, sheet_name=None)
        elif os.path.exists(INPUT_FILE):
            print(f"📄 Loading input data from {INPUT_FILE}...")
            input_data = pd.read_excel(INPUT_FILE, sheet_name=None)
        else:
            print(f"❌ Input file {INPUT_FILE} not found.")
            return
            
        # 2. Extract Target Carriers
        target_carriers = set()
        for df in input_data.values():
            if 'Carrier' in df.columns:
                carriers = df['Carrier'].dropna().unique()
                for c in carriers:
                    c_lower = str(c).lower().strip()
                    target_carriers.add(c_lower)
                    if c_lower in CARRIER_ALIASES:
                        target_carriers.update(CARRIER_ALIASES[c_lower])
        
        target_carriers = list(target_carriers)
        print(f"🎯 Found {len(target_carriers)} unique carrier(s) in Excel. ONLY scanning Drive for these...")
        
        # 3. Cache ONLY the necessary folders
        if self.drive_service:
            self.cache_carrier_statements(target_carriers)
        else:
            print("⚠️ Running without Google Drive validation...")
            
        processed_data = {}
        processed_count = 0
        total_rows = sum(len(df) for df in input_data.values())
        print(f"📊 Total records to process: {total_rows}")
        
        for sheet_name, df in input_data.items():
            print(f"📝 Processing sheet: {sheet_name}")
            
            # Ensure new columns exist
            if 'Drive Match' not in df.columns:
                df['Drive Match'] = ""
                df['Reason'] = ""
                df['APL Website Match'] = ""
                df['APL Match Reason'] = ""
                
            for index, row in df.iterrows():
                # Skip already processed if resuming
                if pd.notna(row.get('Drive Match')) and str(row.get('Drive Match')).strip() != "":
                    if str(row.get('APL Website Match')) != "Error":
                        processed_count += 1
                        continue
                    
                # Assuming 'Processed Policy Number' is the column name
                policy_num = row.get('Processed Policy Number', row.get('Policy Number', ''))
                
                # Fix pandas .0 float formatting bug
                if pd.notna(policy_num):
                    policy_num_str = str(policy_num).strip()
                    if policy_num_str.endswith('.0'):
                        policy_num_str = policy_num_str[:-2]
                    policy_num = policy_num_str
                
                carrier_val = row.get('Carrier')
                actual_carrier = carrier_val if pd.notna(carrier_val) else sheet_name
                
                print(f"  🔍 Validating Policy: {policy_num} for carrier '{actual_carrier}'")
                if self.drive_service:
                    drive_match, drive_reason = self.validate_in_drive(policy_num, actual_carrier)
                else:
                    drive_match, drive_reason = "No Match", "Google Drive API not authenticated"
                
                df.at[index, 'Drive Match'] = drive_match
                df.at[index, 'Reason'] = drive_reason
                
                if drive_match == "Exact Match":
                    apl_match, apl_reason = self.verify_apl_portal(policy_num)
                    df.at[index, 'APL Website Match'] = apl_match
                    df.at[index, 'APL Match Reason'] = apl_reason
                else:
                    df.at[index, 'APL Website Match'] = "Skipped"
                    df.at[index, 'APL Match Reason'] = "No Exact Match in Drive"
                    
                processed_count += 1
                
                # Checkpoint saving
                if processed_count % CHECKPOINT_INTERVAL == 0:
                    print(f"💾 Saving checkpoint at {processed_count}/{total_rows} records...")
                    with pd.ExcelWriter(PROGRESS_FILE, engine='openpyxl') as writer:
                        for s_name, s_df in input_data.items():
                            s_df.to_excel(writer, sheet_name=s_name, index=False)
            
            processed_data[sheet_name] = df
            
        print("✅ Processing complete. Saving final output...")
        
        self.close_browser()
        
        with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
            for s_name, df in processed_data.items():
                df.to_excel(writer, sheet_name=s_name, index=False)
                
        self.format_excel_output(OUTPUT_FILE)
        
        # Remove checkpoint after successful completion
        if os.path.exists(PROGRESS_FILE):
            os.remove(PROGRESS_FILE)
            
        print(f"🎉 Final output saved to {OUTPUT_FILE} and fully formatted!")

if __name__ == "__main__":
    agent = ValidationAgent()
    agent.run()
